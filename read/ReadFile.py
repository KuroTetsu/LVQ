from openpyxl import load_workbook

#read and get data from excel using openpyxl
class OpenFile:
    def GetFile(self, link):
        self.File = load_workbook(link, data_only=True)
        sheet = self.File[self.File.sheetnames[0]]
        self.data = []
        for row in sheet.iter_rows():
            self.bar = []
            for cell in row:
                 self.bar.append(cell.value)
            self.data.append(self.bar)
        return self.data